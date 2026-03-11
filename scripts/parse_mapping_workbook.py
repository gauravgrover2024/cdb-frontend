#!/usr/bin/env python3
"""
Parse edited mapping workbook and emit machine-readable JSON mapping config.
"""

from __future__ import annotations

import argparse
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "migration_analysis/loan_field_mapping_workbook.xlsx"
DEFAULT_OUT = ROOT / "migration_analysis/loan_field_mapping_config.json"


def clean(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _extract_rows_from_sheet(ws):
    headers = {}
    for c in range(1, ws.max_column + 1):
        key = clean(ws.cell(row=1, column=c).value)
        if key:
            headers[key] = c

    required = ["new_field", "legacy_file", "legacy_field", "fallback_priority", "transform", "active", "new_page", "new_section"]
    missing = [x for x in required if x not in headers]
    if missing:
        return []

    rows = []
    for r in range(2, ws.max_row + 1):
        new_field = clean(ws.cell(r, headers["new_field"]).value).strip("`").strip()
        if not new_field:
            continue

        active = clean(ws.cell(r, headers["active"]).value).upper()
        if active in {"N", "NO", "0", "FALSE"}:
            continue

        legacy_file = clean(ws.cell(r, headers["legacy_file"]).value)
        legacy_field = clean(ws.cell(r, headers["legacy_field"]).value)
        if not legacy_file or not legacy_field:
            continue

        new_page = clean(ws.cell(r, headers["new_page"]).value)
        new_section = clean(ws.cell(r, headers["new_section"]).value)
        transform = clean(ws.cell(r, headers["transform"]).value) or "identity"
        notes = clean(ws.cell(r, headers["notes"]).value) if "notes" in headers else ""
        priority_raw = clean(ws.cell(r, headers["fallback_priority"]).value)
        try:
            priority = int(float(priority_raw)) if priority_raw else 1
        except Exception:
            priority = 1

        rows.append(
            {
                "newField": new_field,
                "newPage": new_page,
                "newSection": new_section,
                "legacyFile": legacy_file,
                "legacyField": legacy_field,
                "sourcePath": f"{legacy_file}.{legacy_field}",
                "fallbackPriority": priority,
                "transform": transform,
                "notes": notes,
            }
        )
    return rows


def parse_workbook(workbook_path: Path) -> Dict:
    wb = load_workbook(workbook_path, data_only=True)
    if "MappingEditor" not in wb.sheetnames:
        raise ValueError("Sheet 'MappingEditor' not found")

    raw_rows = []
    # Prefer FallbackEditor when available (clean fallback editing view),
    # then include MappingEditor rows too.
    if "FallbackEditor" in wb.sheetnames:
        raw_rows.extend(_extract_rows_from_sheet(wb["FallbackEditor"]))
    raw_rows.extend(_extract_rows_from_sheet(wb["MappingEditor"]))

    # Deduplicate exact rows coming from both sheets.
    dedup = {}
    for row in raw_rows:
        key = (
            row["newField"],
            row["legacyFile"],
            row["legacyField"],
            row["fallbackPriority"],
            row["transform"],
        )
        dedup[key] = row
    rows_processed = len(dedup)

    grouped = {}
    for row in dedup.values():
        new_field = row["newField"]
        grouped.setdefault(
            new_field,
            {
                "newField": new_field,
                "newPage": row["newPage"],
                "newSection": row["newSection"],
                "legacyMappings": [],
            },
        )
        grouped[new_field]["legacyMappings"].append(
            {
                "legacyFile": row["legacyFile"],
                "legacyField": row["legacyField"],
                "sourcePath": row["sourcePath"],
                "fallbackPriority": row["fallbackPriority"],
                "transform": row["transform"],
                "notes": row["notes"],
            }
        )

    # Sort fallbacks by priority.
    for _, obj in grouped.items():
        obj["legacyMappings"].sort(key=lambda x: (x["fallbackPriority"], x["legacyFile"], x["legacyField"]))

    mappings = sorted(grouped.values(), key=lambda x: x["newField"].lower())
    flat_mappings = []
    for m in mappings:
        for legacy in m["legacyMappings"]:
            flat_mappings.append(
                {
                    "targetField": m["newField"],
                    "newPage": m["newPage"],
                    "newSection": m["newSection"],
                    "legacyFile": legacy["legacyFile"],
                    "legacyField": legacy["legacyField"],
                    "sourcePath": legacy["sourcePath"],
                    "fallbackPriority": legacy["fallbackPriority"],
                    "transform": legacy["transform"],
                    "notes": legacy["notes"],
                }
            )

    return {
        "version": 1,
        "generatedAt": datetime.now(timezone.utc).isoformat(),
        "sourceWorkbook": str(workbook_path),
        "summary": {
            "mappedNewFields": len(mappings),
            "mappingRows": rows_processed,
        },
        "mappingsByNewField": mappings,
        "flatMappings": flat_mappings,
    }


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", default=str(DEFAULT_WORKBOOK))
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    args = parser.parse_args()

    workbook_path = Path(args.workbook)
    out_path = Path(args.out)

    result = parse_workbook(workbook_path)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(json.dumps(result, indent=2), encoding="utf-8")

    print(
        json.dumps(
            {
                "ok": True,
                "workbook": str(workbook_path),
                "out": str(out_path),
                "mappedNewFields": result["summary"]["mappedNewFields"],
                "mappingRows": result["summary"]["mappingRows"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()

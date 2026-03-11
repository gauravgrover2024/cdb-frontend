#!/usr/bin/env python3
"""
Generate an editable Excel workbook for legacy->new loan field mapping.

Features:
- New field catalog with inferred page/section from source files.
- Legacy catalog built from JSON file headers.
- Mapping editor with dropdowns:
  - Legacy File (first dropdown)
  - Legacy Field (dependent dropdown based on selected file)
- Supports multiple fallback mappings by using multiple rows per new field
  with fallback priority.
"""

from __future__ import annotations

import argparse
import json
import os
import re
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SCHEMA = ROOT / "src/modules/loans/schema/loan-module.import.schema.json"
DEFAULT_MAPPING_MD = ROOT / "migration_analysis/legacy_to_new_field_mapping.md"
DEFAULT_OUTPUT = ROOT / "migration_analysis/loan_field_mapping_workbook.xlsx"
DEFAULT_LEGACY_DIR = Path("/Users/gauravgrover/oracle/mil2_json_export")
DEFAULT_EXTRACTED = ROOT / "migration_analysis/pilot_post_output/case_match_extracted.full.json"
DEFAULT_FIELD_CATALOG = ROOT / "migration_analysis/loan_field_catalog.json"

LEGACY_FILES = [
    "AUTH_SIGNATORY.json",
    "CPV_DETAIL.json",
    "CUSTOMER_BANK.json",
    "GURANTOR.json",
    "RC_CUSTOMER_ACCOUNT.json",
    "RC_DOC_DESPATCH_MASTER.json",
    "RC_INSTRUMENT_DETAIL.json",
    "RC_RC_INV_RECEIVING_DETAIL.json",
    "RC_RC_INV_STATUS.json",
    "RC_SOURCE_REFERENCE.json",
]

FLOW_CONDITIONS: Dict[str, Dict[str, str]] = {
    "typeOfLoan": {
        "applies_when": "Always (drives downstream flow)",
        "flow_notes": "Core flow selector: New Car / Used Car / Car Cash-in / Refinance.",
    },
    "isFinanced": {
        "applies_when": "Always",
        "flow_notes": "If No => cash flow (Profile + Delivery only). If Yes => Pre-File/Approval/Post-File/Delivery/Payout flow.",
    },
    "applicantType": {
        "applies_when": "Always",
        "flow_notes": "Individual vs Company changes form visibility and required sections.",
    },
    "currentStage": {
        "applies_when": "Always",
        "flow_notes": "Must respect legacy approval/disbursement progression for migration cases.",
    },
    "approval_status": {
        "applies_when": "Financed flow",
        "flow_notes": "Set approved/disbursed correctly; downstream stages depend on it.",
    },
    "approval_approvalDate": {
        "applies_when": "Financed flow",
        "flow_notes": "For legacy migration, use disbursement date when explicit approval date is absent.",
    },
    "approval_disbursedDate": {
        "applies_when": "Financed flow",
        "flow_notes": "Critical for post-file/delivery unlock in migrated records.",
    },
    "disbursement_date": {
        "applies_when": "Financed flow",
        "flow_notes": "Keep aligned with approval_disbursedDate.",
    },
    "sourceName": {
        "applies_when": "Record source = Indirect (or available source capture)",
        "flow_notes": "Map from SOURCE-like legacy field. Do NOT swap with dealtBy.",
    },
    "dealtBy": {
        "applies_when": "Record details / dispatch handling",
        "flow_notes": "Map from DEALT_BY-like legacy field. Do NOT swap with sourceName.",
    },
    "docsPreparedBy": {
        "applies_when": "Pre-file/Post-file record details",
        "flow_notes": "Map from PRE/POST docs prepared-by chain, not source channel.",
    },
}


def infer_page(source_files: List[str]) -> str:
    text = " ".join(source_files).lower()
    if "/customer-profile/" in text:
        return "Profile"
    if "/pre-file/" in text:
        return "Pre-File"
    if "/loan-approval/" in text:
        return "Approval"
    if "/post-file/" in text:
        return "Post-File"
    if "/vehicle-delivery/" in text:
        return "Delivery"
    if "/payout/" in text:
        return "Payout"
    if "emicalculator" in text or "/emi" in text:
        return "EMI / Utility"
    if "loanstickyheader" in text:
        return "Header / Shared"
    return "Shared"


def is_loan_form_relevant(source_files: List[str]) -> bool:
    return any("/loan-form/" in str(src).replace("\\", "/") for src in (source_files or []))


def normalize_source_files(raw_source_files) -> List[str]:
    out: List[str] = []
    if raw_source_files is None:
        return out
    if not isinstance(raw_source_files, list):
        raw_source_files = [raw_source_files]
    for entry in raw_source_files:
        if entry is None:
            continue
        s = str(entry)
        # Some schema rows store multiple files in one string separated by ';'
        # or commas; split and normalize to one-path-per-item.
        parts = re.split(r"[;,]\s*", s)
        for part in parts:
            part = part.strip()
            if part:
                out.append(part)
    # stable unique
    seen = set()
    uniq = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


def infer_page_from_stage(stage: str) -> str:
    s = str(stage or "").strip().lower()
    if s == "customer-profile":
        return "Profile"
    if s == "pre-file":
        return "Pre-File"
    if s == "loan-approval":
        return "Approval"
    if s == "post-file":
        return "Post-File"
    if s == "vehicle-delivery":
        return "Delivery"
    if s == "payout":
        return "Payout"
    return "Shared"


def infer_section(source_files: List[str]) -> str:
    if not source_files:
        return "Unknown"
    first = source_files[0]
    base = Path(first).stem
    # Convert camel/pascal/snake/kebab to readable.
    base = re.sub(r"([a-z])([A-Z])", r"\1 \2", base)
    base = base.replace("_", " ").replace("-", " ")
    return " ".join(base.split()).title()


def infer_type_from_input_hints(input_types: List[str]) -> str:
    hints = [str(x).lower().strip() for x in (input_types or [])]
    if any(h in {"number"} for h in hints):
        return "number"
    if any(h in {"checkbox", "switch"} for h in hints):
        return "boolean"
    # Keep date/time as string for mapping sheet consistency.
    return "string"


def prettify_field_name(field: str) -> str:
    s = str(field or "").strip()
    if not s:
        return ""
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", s)
    s = s.replace("_", " ").replace("-", " ")
    return " ".join(s.split()).title()


def get_flow_meta(new_field: str) -> Tuple[str, str]:
    meta = FLOW_CONDITIONS.get(new_field, {})
    return meta.get("applies_when", "Always"), meta.get("flow_notes", "")


def load_new_fields(schema_path: Path) -> List[Dict[str, str]]:
    obj = json.loads(schema_path.read_text(encoding="utf-8"))
    props = obj.get("properties", {})
    rows: Dict[str, Dict[str, str]] = {}
    label_by_field: Dict[str, str] = {}
    if DEFAULT_FIELD_CATALOG.exists():
        try:
            cat = json.loads(DEFAULT_FIELD_CATALOG.read_text(encoding="utf-8"))
            for u in cat.get("uniqueFields", []) or []:
                field = str(u.get("field") or "").strip()
                if not field:
                    continue
                label = str(u.get("primaryLabel") or "").strip()
                if label:
                    label_by_field[field] = label
        except Exception:
            pass
    for field, meta in sorted(props.items(), key=lambda x: x[0].lower()):
        source_files = normalize_source_files(meta.get("x-sourceFiles", []) or [])
        source_files = [s for s in source_files if "/loan-form/" in str(s).replace("\\", "/")]
        if not is_loan_form_relevant(source_files):
            continue
        display_label = (
            label_by_field.get(field)
            or str(meta.get("title") or "").strip()
            or prettify_field_name(field)
        )
        rows[field] = {
            "new_field": field,
            "display_label": display_label,
            "new_field_display": f"{field} ({display_label})",
            "new_type": str(meta.get("type", "")),
            "new_page": infer_page(source_files),
            "new_section": infer_section(source_files),
            "new_source_files": "; ".join(source_files),
        }

    # Augment with fields discovered directly from form components (for newly added fields
    # that may not yet exist in schema files).
    if DEFAULT_FIELD_CATALOG.exists():
        try:
            cat = json.loads(DEFAULT_FIELD_CATALOG.read_text(encoding="utf-8"))
            for u in cat.get("uniqueFields", []) or []:
                field = str(u.get("field") or "").strip()
                if not field or field.startswith("__"):
                    continue
                if field in rows:
                    # Enrich source files if schema list is sparse.
                    if not rows[field].get("new_source_files"):
                        files = normalize_source_files(u.get("files") or [])
                        files = [s for s in files if "/loan-form/" in str(s).replace("\\", "/")]
                        rows[field]["new_source_files"] = "; ".join(files)
                    continue
                files = normalize_source_files(u.get("files") or [])
                files = [s for s in files if "/loan-form/" in str(s).replace("\\", "/")]
                if not is_loan_form_relevant(files):
                    continue
                stages = [str(x) for x in (u.get("stages") or []) if x]
                components = [str(x) for x in (u.get("components") or []) if x]
                input_types = [str(x) for x in (u.get("inputTypes") or []) if x]
                display_label = str(u.get("primaryLabel") or "").strip() or prettify_field_name(field)
                rows[field] = {
                    "new_field": field,
                    "display_label": display_label,
                    "new_field_display": f"{field} ({display_label})",
                    "new_type": infer_type_from_input_hints(input_types),
                    "new_page": infer_page_from_stage(stages[0] if stages else ""),
                    "new_section": (components[0] if components else "Unknown"),
                    "new_source_files": "; ".join(files),
                }
        except Exception:
            pass

    return [rows[k] for k in sorted(rows.keys(), key=lambda x: x.lower())]


def _extract_source_paths(expr: str) -> List[str]:
    """
    Extract FILE.FIELD entries from markdown mapping text.
    """
    if not expr:
        return []
    files = [f.replace(".json", "") for f in LEGACY_FILES]
    file_group = "|".join(re.escape(f) for f in files)
    pattern = re.compile(rf"\b({file_group})\.([A-Za-z0-9_]+)\b")
    seen = []
    for m in pattern.finditer(expr):
        path = f"{m.group(1)}.json.{m.group(2)}"
        if path not in seen:
            seen.append(path)
    return seen


def load_prefill_mapping_rows(mapping_md: Path) -> List[Dict[str, str]]:
    """
    Parse markdown tables with columns:
    | New field | Legacy source(s) | Transform / rule |
    Returns one row per fallback source path.
    """
    if not mapping_md.exists():
        return []
    lines = mapping_md.read_text(encoding="utf-8").splitlines()
    rows: List[Dict[str, str]] = []
    in_table = False
    for line in lines:
        s = line.strip()
        if s.startswith("| New field | Legacy source(s) | Transform / rule |"):
            in_table = True
            continue
        if in_table and s.startswith("|---|---|---|"):
            continue
        if in_table and s.startswith("|") and s.endswith("|"):
            parts = [p.strip() for p in s.split("|")[1:-1]]
            if len(parts) < 3:
                continue
            new_field, legacy_expr, transform = parts[0], parts[1], parts[2]
            new_field = new_field.strip().strip("`").strip()
            source_paths = _extract_source_paths(legacy_expr)
            if not source_paths:
                continue
            for idx, sp in enumerate(source_paths, start=1):
                legacy_file, legacy_field = sp.split(".json.", 1)
                rows.append(
                    {
                        "new_field": new_field,
                        "legacy_file": f"{legacy_file}.json",
                        "legacy_field": legacy_field,
                        "fallback_priority": idx,
                        "transform": "identity" if "derived" not in transform.lower() else "derived",
                        "notes": transform,
                    }
                )
            continue
        if in_table and (not s or not s.startswith("|")):
            in_table = False
    return rows


def load_legacy_fields(legacy_dir: Path) -> Dict[str, List[str]]:
    catalog: Dict[str, List[str]] = {}
    for file_name in LEGACY_FILES:
        p = legacy_dir / file_name
        if not p.exists():
            catalog[file_name] = []
            continue
        try:
            arr = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            catalog[file_name] = []
            continue
        fields = set()
        if isinstance(arr, list):
            for row in arr[:500]:
                if isinstance(row, dict):
                    fields.update(row.keys())
        catalog[file_name] = sorted(fields, key=lambda x: x.lower())
    return catalog


def sanitize_named_range(file_name: str) -> str:
    base = file_name.replace(".json", "")
    base = re.sub(r"[^A-Za-z0-9_]", "_", base)
    if re.match(r"^[0-9]", base):
        base = f"F_{base}"
    return base


def autosize(ws, min_width: int = 12, max_width: int = 60):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def style_header(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def stringify_value(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def first_meaningful(values):
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return v
    return None


def load_case_source_map(extracted_path: Path, case_id: str) -> Dict[str, str]:
    if not case_id or not extracted_path.exists():
        return {}
    try:
        obj = json.loads(extracted_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    case_obj = obj.get(str(case_id))
    if not case_obj:
        return {}
    matches = case_obj.get("matches", {}) or {}
    out: Dict[str, str] = {}
    for file_name, rows in matches.items():
        if not isinstance(rows, list):
            continue
        # Gather all field names present in any row.
        field_names = set()
        for row in rows:
            if isinstance(row, dict):
                field_names.update(row.keys())
        for field in sorted(field_names):
            vals = []
            for row in rows:
                if isinstance(row, dict) and field in row:
                    vals.append(row.get(field))
            v = first_meaningful(vals)
            if v is None and vals:
                v = vals[0]
            out[f"{file_name}.{field}"] = stringify_value(v)
    return out


def build_workbook(
    output_path: Path,
    new_fields: List[Dict[str, str]],
    legacy_catalog: Dict[str, List[str]],
    prefill_rows: List[Dict[str, str]],
    case_id: str = "",
    case_source_map: Dict[str, str] | None = None,
):
    wb = Workbook()
    ws_map = wb.active
    ws_map.title = "MappingEditor"
    ws_new = wb.create_sheet("NewFieldCatalog")
    ws_legacy = wb.create_sheet("LegacyCatalog")
    ws_dd = wb.create_sheet("DropdownData")
    ws_instr = wb.create_sheet("Instructions")
    ws_unmapped_new = wb.create_sheet("UnmappedNewFields")
    ws_unmapped_legacy = wb.create_sheet("UnmappedLegacyFields")
    ws_fallback = wb.create_sheet("FallbackEditor")
    ws_case_source = wb.create_sheet("CaseSource")

    # New field catalog
    ws_new.append([
        "new_field",
        "new_field_display",
        "display_label",
        "new_type",
        "new_page",
        "new_section",
        "applies_when",
        "flow_notes",
        "new_source_files",
    ])
    for row in new_fields:
        applies_when, flow_notes = get_flow_meta(row["new_field"])
        ws_new.append(
            [
                row["new_field"],
                row.get("new_field_display", row["new_field"]),
                row.get("display_label", ""),
                row["new_type"],
                row["new_page"],
                row["new_section"],
                applies_when,
                flow_notes,
                row["new_source_files"],
            ]
        )
    style_header(ws_new)
    autosize(ws_new)

    # Legacy catalog
    ws_legacy.append(["legacy_file", "legacy_field"])
    for file_name, fields in legacy_catalog.items():
        if not fields:
            ws_legacy.append([file_name, ""])
        else:
            for field in fields:
                ws_legacy.append([file_name, field])
    style_header(ws_legacy)
    autosize(ws_legacy)

    # Dropdown data sheet:
    # A: legacy file list
    # B..: each file fields in separate column + named range
    ws_dd["A1"] = "legacy_files"
    row = 2
    file_names = sorted(legacy_catalog.keys())
    for f in file_names:
        ws_dd.cell(row=row, column=1, value=f)
        row += 1

    # Define named range for all legacy files
    wb.defined_names.add(
        DefinedName(
            "LegacyFiles",
            attr_text=f"'DropdownData'!$A$2:$A${1 + len(file_names)}",
        )
    )

    col = 2
    for file_name in file_names:
        fields = legacy_catalog.get(file_name, [])
        if not fields:
            fields = [""]
        ws_dd.cell(row=1, column=col, value=file_name)
        for i, field in enumerate(fields, start=2):
            ws_dd.cell(row=i, column=col, value=field)
        col_letter = ws_dd.cell(row=1, column=col).column_letter
        range_name = sanitize_named_range(file_name)
        wb.defined_names.add(
            DefinedName(
                range_name,
                attr_text=f"'DropdownData'!${col_letter}$2:${col_letter}${1 + len(fields)}",
            )
        )
        col += 1

    # Mapping editor
    ws_map.append(
        [
            "new_field",
            "new_page",
            "new_section",
            "applies_when",
            "flow_notes",
            "legacy_file",
            "legacy_field",
            "fallback_priority",
            "transform",
            "active",
            "notes",
            "example_legacy_value",
            "example_new_value",
            "case_value_preview",
            "new_field_display",
        ]
    )

    field_meta = {r["new_field"]: r for r in new_fields}
    allowed_new_fields = set(field_meta.keys())
    prefill_rows = [r for r in prefill_rows if r.get("new_field") in allowed_new_fields]

    # Write prefilled mappings first (including fallback rows).
    for pr in prefill_rows:
        meta = field_meta.get(pr["new_field"], {})
        applies_when, flow_notes = get_flow_meta(pr["new_field"])
        ws_map.append(
            [
                pr["new_field"],
                meta.get("new_page", ""),
                meta.get("new_section", ""),
                applies_when,
                flow_notes,
                pr["legacy_file"],
                pr["legacy_field"],
                pr["fallback_priority"],
                pr.get("transform", "identity"),
                "Y",
                pr.get("notes", ""),
                "",
                "",
                "",
                meta.get("new_field_display", pr["new_field"]),
            ]
        )

    # Add one blank editable row for unmapped new fields.
    prefilled_fields = {r["new_field"] for r in prefill_rows}
    for row in new_fields:
        if row["new_field"] in prefilled_fields:
            continue
        applies_when, flow_notes = get_flow_meta(row["new_field"])
        ws_map.append(
            [
                row["new_field"],
                row["new_page"],
                row["new_section"],
                applies_when,
                flow_notes,
                "",
                "",
                1,
                "identity",
                "Y",
                "",
                "",
                "",
                "",
                row.get("new_field_display", row["new_field"]),
            ]
        )

    # Build coverage indexes
    mapped_by_field: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    used_legacy_pairs = set()
    used_legacy_by_file: Dict[str, set] = defaultdict(set)
    for r in prefill_rows:
        mapped_by_field[r["new_field"]].append(r)
        used_legacy_pairs.add((r["legacy_file"], r["legacy_field"]))
        used_legacy_by_file[r["legacy_file"]].add(r["legacy_field"])
    style_header(ws_map)

    # Add validations for a broad editable range.
    total_rows = max(5000, len(new_fields) + 2000)
    # Legacy file dropdown
    dv_file = DataValidation(type="list", formula1="=LegacyFiles", allow_blank=True)
    ws_map.add_data_validation(dv_file)
    dv_file.add(f"F2:F{total_rows}")

    # Legacy field dependent dropdown (depends on column D value in same row).
    # Convert "RC_CUSTOMER_ACCOUNT.json" -> "RC_CUSTOMER_ACCOUNT"
    dv_field = DataValidation(
        type="list",
        formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE($F2,".json","")," ","_"))',
        allow_blank=True,
    )
    ws_map.add_data_validation(dv_field)
    dv_field.add(f"G2:G{total_rows}")

    autosize(ws_map)
    ws_map.column_dimensions["D"].width = 28
    ws_map.column_dimensions["E"].width = 56
    ws_map.column_dimensions["F"].width = 28
    ws_map.column_dimensions["G"].width = 36
    ws_map.column_dimensions["K"].width = 50
    ws_map.column_dimensions["L"].width = 30
    ws_map.column_dimensions["M"].width = 30
    ws_map.column_dimensions["N"].width = 36
    ws_map.column_dimensions["O"].width = 44

    # Instructions
    ws_instr["A1"] = "How to use this workbook"
    ws_instr["A1"].font = Font(bold=True, size=14)
    instructions = [
        "1. Open sheet 'MappingEditor'.",
        "2. Columns 'applies_when' and 'flow_notes' explain conditional flow dependencies.",
        "3. Each row is one mapping rule. Duplicate rows to add fallbacks for the same new_field.",
        "4. Select 'legacy_file' from dropdown first, then 'legacy_field' from dependent dropdown.",
        "5. Use fallback_priority: 1 (highest), 2, 3 ...",
        "6. Set transform (identity, normalize_yes_no, to_date, to_number, concat, custom).",
        "7. Set active = Y/N.",
        "8. Important: keep Source vs Dealt By separate (sourceName <- SOURCE-like, dealtBy <- DEALT_BY-like).",
        "9. Save workbook and run parser script to generate JSON mapping config.",
    ]
    for idx, line in enumerate(instructions, start=3):
        ws_instr[f"A{idx}"] = line
    ws_instr["A13"] = "Parser command:"
    ws_instr["A14"] = (
        "python3 scripts/parse_mapping_workbook.py "
        "--workbook migration_analysis/loan_field_mapping_workbook.xlsx "
        "--out migration_analysis/loan_field_mapping_config.json"
    )
    ws_instr["A16"] = "Auto status:"
    ws_instr["A17"] = "Unmapped sheets recalculate automatically from MappingEditor/FallbackEditor."
    ws_instr["A18"] = "Case preview:"
    ws_instr["A19"] = f"Workbook generated for case: {case_id or '(none)'}"
    ws_instr.column_dimensions["A"].width = 140

    # Unmapped new fields sheet
    ws_unmapped_new.append(["new_field", "new_page", "new_section", "status"])
    for idx, row in enumerate(new_fields, start=2):
        ws_unmapped_new.append([row["new_field"], row["new_page"], row["new_section"], ""])
        ws_unmapped_new[f"D{idx}"] = (
            f'=IF((COUNTIFS(MappingEditor!$A:$A,$A{idx},MappingEditor!$J:$J,"Y")'
            f'+COUNTIFS(FallbackEditor!$A:$A,$A{idx},FallbackEditor!$J:$J,"Y"))>0,"mapped","unmapped")'
        )
    style_header(ws_unmapped_new)
    autosize(ws_unmapped_new)

    # Unmapped legacy fields sheet
    ws_unmapped_legacy.append(["legacy_file", "legacy_field", "status"])
    row_idx = 2
    for file_name, fields in legacy_catalog.items():
        for field in fields:
            ws_unmapped_legacy.append([file_name, field, ""])
            ws_unmapped_legacy[f"C{row_idx}"] = (
                f'=IF((COUNTIFS(MappingEditor!$F:$F,$A{row_idx},MappingEditor!$G:$G,$B{row_idx},MappingEditor!$J:$J,"Y")'
                f'+COUNTIFS(FallbackEditor!$G:$G,$A{row_idx},FallbackEditor!$H:$H,$B{row_idx},FallbackEditor!$J:$J,"Y"))>0,"mapped","unmapped")'
            )
            row_idx += 1
    style_header(ws_unmapped_legacy)
    autosize(ws_unmapped_legacy)

    # Fallback editor sheet (clean view for easy fallback changes)
    ws_fallback.append(
        [
            "new_field",
            "new_page",
            "new_section",
            "applies_when",
            "flow_notes",
            "fallback_priority",
            "legacy_file",
            "legacy_field",
            "transform",
            "active",
            "notes",
            "case_value_preview",
            "new_field_display",
        ]
    )
    for new_field in sorted(mapped_by_field.keys(), key=lambda x: x.lower()):
        meta = field_meta.get(new_field, {})
        rules = sorted(
            mapped_by_field[new_field],
            key=lambda x: int(x.get("fallback_priority") or 1),
        )
        for rule in rules:
            applies_when, flow_notes = get_flow_meta(new_field)
            ws_fallback.append(
                [
                    new_field,
                    meta.get("new_page", ""),
                    meta.get("new_section", ""),
                    applies_when,
                    flow_notes,
                    rule.get("fallback_priority", 1),
                    rule.get("legacy_file", ""),
                    rule.get("legacy_field", ""),
                    rule.get("transform", "identity"),
                    "Y",
                    rule.get("notes", ""),
                    "",
                    meta.get("new_field_display", new_field),
                ]
            )
    style_header(ws_fallback)
    autosize(ws_fallback)
    ws_fallback.column_dimensions["D"].width = 28
    ws_fallback.column_dimensions["E"].width = 56
    ws_fallback.column_dimensions["F"].width = 16
    ws_fallback.column_dimensions["G"].width = 28
    ws_fallback.column_dimensions["H"].width = 36
    ws_fallback.column_dimensions["K"].width = 50
    ws_fallback.column_dimensions["L"].width = 36
    ws_fallback.column_dimensions["M"].width = 44

    fb_total_rows = max(5000, ws_fallback.max_row + 2000)
    fb_dv_file = DataValidation(type="list", formula1="=LegacyFiles", allow_blank=True)
    ws_fallback.add_data_validation(fb_dv_file)
    fb_dv_file.add(f"G2:G{fb_total_rows}")
    fb_dv_field = DataValidation(
        type="list",
        formula1='=INDIRECT(SUBSTITUTE(SUBSTITUTE($G2,".json","")," ","_"))',
        allow_blank=True,
    )
    ws_fallback.add_data_validation(fb_dv_field)
    fb_dv_field.add(f"H2:H{fb_total_rows}")

    # Case source sheet and preview formulas
    case_source_map = case_source_map or {}
    ws_case_source.append(["source_path", "value"])
    for key in sorted(case_source_map.keys(), key=lambda x: x.lower()):
        ws_case_source.append([key, case_source_map[key]])
    style_header(ws_case_source)
    autosize(ws_case_source, max_width=80)

    if ws_case_source.max_row >= 2:
        source_range = f"CaseSource!$A$2:$A${ws_case_source.max_row}"
        value_range = f"CaseSource!$B$2:$B${ws_case_source.max_row}"
        for r in range(2, ws_map.max_row + 1):
            ws_map[f"N{r}"] = f'=IFERROR(XLOOKUP($F{r}&"."&$G{r},{source_range},{value_range},""),"")'
        for r in range(2, ws_fallback.max_row + 1):
            ws_fallback[f"L{r}"] = f'=IFERROR(XLOOKUP($G{r}&"."&$H{r},{source_range},{value_range},""),"")'
    else:
        for r in range(2, ws_map.max_row + 1):
            ws_map[f"N{r}"] = ""
        for r in range(2, ws_fallback.max_row + 1):
            ws_fallback[f"L{r}"] = ""

    # Hide helper sheet
    ws_dd.sheet_state = "hidden"
    style_header(ws_map)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--schema", default=str(DEFAULT_SCHEMA))
    parser.add_argument("--legacy-dir", default=str(DEFAULT_LEGACY_DIR))
    parser.add_argument("--out", default=str(DEFAULT_OUTPUT))
    parser.add_argument("--case-id", default="")
    parser.add_argument("--extracted", default=str(DEFAULT_EXTRACTED))
    args = parser.parse_args()

    schema_path = Path(args.schema)
    legacy_dir = Path(args.legacy_dir)
    out_path = Path(args.out)
    extracted_path = Path(args.extracted)
    case_id = str(args.case_id or "").strip()

    new_fields = load_new_fields(schema_path)
    legacy_catalog = load_legacy_fields(legacy_dir)
    prefill_rows = load_prefill_mapping_rows(DEFAULT_MAPPING_MD)
    case_source_map = load_case_source_map(extracted_path, case_id) if case_id else {}
    build_workbook(out_path, new_fields, legacy_catalog, prefill_rows, case_id=case_id, case_source_map=case_source_map)
    print(
        json.dumps(
            {
                "ok": True,
                "generatedAt": datetime.now(timezone.utc).isoformat(),
                "workbook": str(out_path),
                "newFields": len(new_fields),
                "legacyFiles": len(legacy_catalog),
                "prefilledRows": len(prefill_rows),
                "caseId": case_id,
                "caseSourcePaths": len(case_source_map),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()

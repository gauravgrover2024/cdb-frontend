#!/usr/bin/env python3
"""
Generate reverse mapping workbook: legacy -> new fields (case-specific).

This workbook is intentionally Excel-compatible without relying on dynamic
functions like FILTER/XLOOKUP. It supports multi-target mapping by providing
5 new-field dropdown slots per legacy field row.
"""

from __future__ import annotations

import argparse
import json
import re
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.defined_name import DefinedName


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_SCHEMA = ROOT / "src/modules/loans/schema/loan-module.schema.json"
DEFAULT_CATALOG = ROOT / "migration_analysis/loan_field_catalog.json"
DEFAULT_EXTRACTED = ROOT / "migration_analysis/pilot_post_output/case_match_extracted.full.json"
DEFAULT_LEGACY_DIR = Path("/Users/gauravgrover/oracle/mil2_json_export")
DEFAULT_OUT = ROOT / "migration_analysis/reverse_mapping_workbook.xlsx"

LEGACY_FILES = [
    "AUTH_SIGNATORY.json",
    "CPV_DETAIL.json",
    "CUSTOMER_BANK.json",
    "GURANTOR.json",
    "RC_CUSTOMER_ACCOUNT.json",
    "RC_DOC_DESPATCH_MASTER.json",
    "RC_INSTRUMENT_DETAIL.json",
    "RC_RC_INV_RECEIVING_DETAIL.json",
    "RC_RC_INV_STATUS.json",
    "RC_SOURCE_REFERENCE.json",
]

# Conditional naming hints for easier identification.
CONDITIONAL_NAME_HINTS = {
    "dob": "Individual: Date of Birth | Company: Date of Incorporation",
    "customerName": "Individual: Customer Name | Company: Company Name",
    "primaryMobile": "Individual: Mobile | Company: Office Mobile",
    "residenceAddress": "Individual: Residence Address | Company: Office Address",
    "pincode": "Individual: Residence Pincode | Company: Office Pincode",
    "city": "Individual: Residence City | Company: Office City",
}


def style_header(ws, row: int = 1):
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    for cell in ws[row]:
        cell.fill = fill
        cell.font = font


def autosize(ws, min_width: int = 10, max_width: int = 70):
    for col in ws.columns:
        letter = col[0].column_letter
        max_len = 0
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            if len(value) > max_len:
                max_len = len(value)
        ws.column_dimensions[letter].width = max(min_width, min(max_width, max_len + 2))


def normalize_source_files(raw_source_files) -> List[str]:
    out: List[str] = []
    if raw_source_files is None:
        return out
    if not isinstance(raw_source_files, list):
        raw_source_files = [raw_source_files]
    for entry in raw_source_files:
        if entry is None:
            continue
        for part in re.split(r"[;,]\s*", str(entry)):
            p = part.strip()
            if p:
                out.append(p)
    # stable unique
    seen = set()
    uniq = []
    for x in out:
        if x in seen:
            continue
        seen.add(x)
        uniq.append(x)
    return uniq


def infer_type_from_input_hints(input_types: List[str]) -> str:
    hints = [str(x).lower().strip() for x in (input_types or [])]
    if any(h in {"number"} for h in hints):
        return "number"
    if any(h in {"checkbox", "switch"} for h in hints):
        return "boolean"
    return "string"


def prettify_field_name(field: str) -> str:
    s = str(field or "").strip()
    if not s:
        return ""
    s = re.sub(r"([a-z0-9])([A-Z])", r"\1 \2", s)
    s = s.replace("_", " ").replace("-", " ")
    return " ".join(s.split()).title()


def is_loan_form_relevant(source_files: List[str]) -> bool:
    return any("/loan-form/" in str(src).replace("\\", "/") for src in (source_files or []))


def load_new_field_options(schema_path: Path, catalog_path: Path) -> List[Dict[str, str]]:
    schema = json.loads(schema_path.read_text(encoding="utf-8"))
    props = schema.get("properties", {})

    label_by_field: Dict[str, str] = {}
    options_by_field: Dict[str, List[str]] = {}
    if catalog_path.exists():
        try:
            cat = json.loads(catalog_path.read_text(encoding="utf-8"))
            for u in cat.get("uniqueFields", []) or []:
                field = str(u.get("field") or "").strip()
                if not field:
                    continue
                lbl = str(u.get("primaryLabel") or "").strip()
                if lbl:
                    label_by_field[field] = lbl
                opts = [str(x) for x in (u.get("options") or []) if str(x).strip()]
                if opts:
                    options_by_field[field] = opts
        except Exception:
            pass

    rows: List[Dict[str, str]] = []
    # schema-backed (loan-form only)
    for field, meta in sorted(props.items(), key=lambda x: x[0].lower()):
        source_files = normalize_source_files(meta.get("x-sourceFiles", []))
        source_files = [s for s in source_files if "/loan-form/" in str(s).replace("\\", "/")]
        if not is_loan_form_relevant(source_files):
            continue
        display_label = (
            label_by_field.get(field)
            or str(meta.get("title") or "").strip()
            or prettify_field_name(field)
        )
        if field in CONDITIONAL_NAME_HINTS:
            display_label = f"{display_label} [{CONDITIONAL_NAME_HINTS[field]}]"
        dropdown_values = " | ".join(options_by_field.get(field, []))
        rows.append(
            {
                "new_field": field,
                "display_label": display_label,
                "display": f"{field} ({display_label})",
                "new_type": str(meta.get("type", "")),
                "dropdown_values": dropdown_values,
            }
        )

    # catalog-only fields not in schema yet (loan-form only)
    existing = {r["new_field"] for r in rows}
    if catalog_path.exists():
        try:
            cat = json.loads(catalog_path.read_text(encoding="utf-8"))
            for u in cat.get("uniqueFields", []) or []:
                field = str(u.get("field") or "").strip()
                if not field or field.startswith("__") or field in existing:
                    continue
                files = normalize_source_files(u.get("files") or [])
                files = [s for s in files if "/loan-form/" in str(s).replace("\\", "/")]
                if not is_loan_form_relevant(files):
                    continue
                display_label = str(u.get("primaryLabel") or "").strip() or prettify_field_name(field)
                if field in CONDITIONAL_NAME_HINTS:
                    display_label = f"{display_label} [{CONDITIONAL_NAME_HINTS[field]}]"
                dropdown_values = " | ".join([str(x) for x in (u.get("options") or []) if str(x).strip()])
                rows.append(
                    {
                        "new_field": field,
                        "display_label": display_label,
                        "display": f"{field} ({display_label})",
                        "new_type": infer_type_from_input_hints([str(x) for x in (u.get("inputTypes") or [])]),
                        "dropdown_values": dropdown_values,
                    }
                )
        except Exception:
            pass

    return sorted(rows, key=lambda r: r["new_field"].lower())


def load_legacy_fields(legacy_dir: Path) -> Dict[str, List[str]]:
    catalog: Dict[str, List[str]] = {}
    for file_name in LEGACY_FILES:
        p = legacy_dir / file_name
        if not p.exists():
            catalog[file_name] = []
            continue
        try:
            arr = json.loads(p.read_text(encoding="utf-8"))
        except Exception:
            catalog[file_name] = []
            continue
        fields = set()
        if isinstance(arr, list):
            for row in arr[:500]:
                if isinstance(row, dict):
                    fields.update(row.keys())
        catalog[file_name] = sorted(fields, key=lambda x: x.lower())
    return catalog


def stringify_value(v):
    if v is None:
        return ""
    if isinstance(v, (dict, list)):
        return json.dumps(v, ensure_ascii=False)
    return str(v)


def first_meaningful(values):
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return v
    return None


def load_case_source_map(extracted_path: Path, case_id: str) -> Dict[str, str]:
    if not case_id or not extracted_path.exists():
        return {}
    try:
        obj = json.loads(extracted_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    case_obj = obj.get(str(case_id))
    if not case_obj:
        return {}
    matches = case_obj.get("matches", {}) or {}
    out: Dict[str, str] = {}
    for file_name, rows in matches.items():
        if not isinstance(rows, list):
            continue
        field_names = set()
        for row in rows:
            if isinstance(row, dict):
                field_names.update(row.keys())
        for field in sorted(field_names):
            vals = []
            for row in rows:
                if isinstance(row, dict) and field in row:
                    vals.append(row.get(field))
            v = first_meaningful(vals)
            if v is None and vals:
                v = vals[0]
            out[f"{file_name}.{field}"] = stringify_value(v)
    return out


def build_workbook(
    output_path: Path,
    new_fields: List[Dict[str, str]],
    legacy_catalog: Dict[str, List[str]],
    case_source_map: Dict[str, str],
):
    wb = Workbook()
    ws_main = wb.active
    ws_main.title = "LegacyToNewMap"
    ws_new = wb.create_sheet("NewFieldOptions")
    ws_dd = wb.create_sheet("DropdownData")
    ws_unmapped_old = wb.create_sheet("UnmappedLegacyFields")
    ws_unmapped_new = wb.create_sheet("UnmappedNewFields")
    ws_instr = wb.create_sheet("Instructions")

    # New fields option sheet
    ws_new.append(["display", "new_field", "display_label", "dropdown_values", "new_type"])
    for row in new_fields:
        ws_new.append(
            [
                row["display"],
                row["new_field"],
                row.get("display_label", ""),
                row.get("dropdown_values", ""),
                row.get("new_type", ""),
            ]
        )
    style_header(ws_new)
    autosize(ws_new)

    # Dropdown helper
    ws_dd["A1"] = "new_field_display_options"
    for i, row in enumerate(new_fields, start=2):
        ws_dd.cell(row=i, column=1, value=row["display"])
    wb.defined_names.add(
        DefinedName(
            "NewFieldOptionsDisplay",
            attr_text=f"'DropdownData'!$A$2:$A${1+len(new_fields)}",
        )
    )

    # Role dropdown lists
    role_first = ["Mapping"] + [f"Fallback{n}" for n in range(1, 21)]
    role_next = [f"Fallback{n}" for n in range(1, 21)]
    ws_dd["C1"] = "role_first"
    ws_dd["D1"] = "role_next"
    for i, v in enumerate(role_first, start=2):
        ws_dd.cell(row=i, column=3, value=v)
    for i, v in enumerate(role_next, start=2):
        ws_dd.cell(row=i, column=4, value=v)
    wb.defined_names.add(
        DefinedName("RoleFirst", attr_text=f"'DropdownData'!$C$2:$C${1+len(role_first)}")
    )
    wb.defined_names.add(
        DefinedName("RoleNext", attr_text=f"'DropdownData'!$D$2:$D${1+len(role_next)}")
    )

    # Main sheet (multi-target mapping via 5 dropdown slots)
    ws_main.append(
        [
            "legacy_file",
            "legacy_field",
            "legacy_value",
            "new_field_1",
            "new_field_2",
            "new_field_3",
            "new_field_4",
            "new_field_5",
            "role_1",
            "role_2",
            "role_3",
            "role_4",
            "role_5",
            "comments_normalization",
            "legacy_map_status",
        ]
    )
    for file_name in LEGACY_FILES:
        fields = legacy_catalog.get(file_name, []) or []
        for fld in fields:
            key = f"{file_name}.{fld}"
            ws_main.append(
                [
                    file_name,
                    fld,
                    case_source_map.get(key, ""),
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                ]
            )
    style_header(ws_main)
    autosize(ws_main)
    ws_main.column_dimensions["A"].width = 24
    ws_main.column_dimensions["B"].width = 28
    ws_main.column_dimensions["C"].width = 36
    ws_main.column_dimensions["D"].width = 36
    ws_main.column_dimensions["E"].width = 36
    ws_main.column_dimensions["F"].width = 36
    ws_main.column_dimensions["G"].width = 36
    ws_main.column_dimensions["H"].width = 36
    ws_main.column_dimensions["I"].width = 14
    ws_main.column_dimensions["J"].width = 14
    ws_main.column_dimensions["K"].width = 14
    ws_main.column_dimensions["L"].width = 14
    ws_main.column_dimensions["M"].width = 14
    ws_main.column_dimensions["N"].width = 42
    ws_main.column_dimensions["O"].width = 18

    total_rows = max(5000, ws_main.max_row + 2000)
    # New field dropdown on all 5 slots
    dv_new = DataValidation(type="list", formula1="=NewFieldOptionsDisplay", allow_blank=True)
    ws_main.add_data_validation(dv_new)
    dv_new.add(f"D2:H{total_rows}")

    # Role dropdown for each slot
    dv_role = DataValidation(type="list", formula1="=RoleFirst", allow_blank=True)
    ws_main.add_data_validation(dv_role)
    dv_role.add(f"I2:M{total_rows}")

    # Status formulas (legacy row mapped/unmapped)
    for r in range(2, ws_main.max_row + 1):
        ws_main[f"O{r}"] = f'=IF(COUNTA(D{r}:H{r})=0,"UNMAPPED","MAPPED")'

    # Unmapped old fields (table + status formula, user can filter status)
    ws_unmapped_old["A1"] = "Legacy field mapping status (filter legacy_map_status = UNMAPPED)"
    ws_unmapped_old.append(["legacy_file", "legacy_field", "legacy_value", "legacy_map_status"])
    for row_idx in range(2, ws_main.max_row + 1):
        ws_unmapped_old.append(
            [
                ws_main[f"A{row_idx}"].value,
                ws_main[f"B{row_idx}"].value,
                ws_main[f"C{row_idx}"].value,
                f"=LegacyToNewMap!O{row_idx}",
            ]
        )
    style_header(ws_unmapped_old, row=2)
    ws_unmapped_old.auto_filter.ref = f"A2:D{ws_unmapped_old.max_row}"
    autosize(ws_unmapped_old)
    ws_unmapped_old.column_dimensions["A"].width = 24
    ws_unmapped_old.column_dimensions["B"].width = 28
    ws_unmapped_old.column_dimensions["C"].width = 36
    ws_unmapped_old.column_dimensions["D"].width = 18

    # Unmapped new fields as a robust table with count formula (no FILTER/XLOOKUP)
    ws_unmapped_new["A1"] = "New field mapping status (filter mapped_count = 0)"
    ws_unmapped_new.append(["new_field_display", "new_field_key", "dropdown_values", "mapped_count"])
    for i, row in enumerate(new_fields, start=3):
        ws_unmapped_new.append(
            [
                row["display"],
                row["new_field"],
                row.get("dropdown_values", ""),
                f"=COUNTIF(LegacyToNewMap!$D:$H,A{i})",
            ]
        )
    style_header(ws_unmapped_new, row=2)
    ws_unmapped_new.auto_filter.ref = f"A2:D{ws_unmapped_new.max_row}"
    autosize(ws_unmapped_new)
    ws_unmapped_new.column_dimensions["A"].width = 50
    ws_unmapped_new.column_dimensions["B"].width = 26
    ws_unmapped_new.column_dimensions["C"].width = 44
    ws_unmapped_new.column_dimensions["D"].width = 14

    # Instructions
    ws_instr["A1"] = "Reverse Mapping Instructions"
    ws_instr["A1"].font = Font(bold=True, size=14)
    lines = [
        "1) Map each legacy field to one or more new fields via new_field_1..new_field_5.",
        "2) Use role_1..role_5 as Mapping/Fallback1/Fallback2 as needed.",
        "3) Same legacy field can map to multiple new fields in the same row.",
        "4) Add normalization rules in comments_normalization (for dropdown conversion etc.).",
        "5) In UnmappedLegacyFields sheet, filter legacy_map_status = UNMAPPED.",
        "6) In UnmappedNewFields sheet, filter mapped_count = 0.",
        "7) No dynamic Excel-only functions are used; workbook is compatibility-safe.",
    ]
    for i, t in enumerate(lines, start=3):
        ws_instr[f"A{i}"] = t
    ws_instr.column_dimensions["A"].width = 140

    # hide helper
    ws_dd.sheet_state = "hidden"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--schema", default=str(DEFAULT_SCHEMA))
    parser.add_argument("--catalog", default=str(DEFAULT_CATALOG))
    parser.add_argument("--legacy-dir", default=str(DEFAULT_LEGACY_DIR))
    parser.add_argument("--extracted", default=str(DEFAULT_EXTRACTED))
    parser.add_argument("--case-id", required=True)
    parser.add_argument("--out", default=str(DEFAULT_OUT))
    args = parser.parse_args()

    schema_path = Path(args.schema)
    catalog_path = Path(args.catalog)
    legacy_dir = Path(args.legacy_dir)
    extracted_path = Path(args.extracted)
    case_id = str(args.case_id).strip()
    out_path = Path(args.out)

    new_fields = load_new_field_options(schema_path, catalog_path)
    legacy_catalog = load_legacy_fields(legacy_dir)
    case_source_map = load_case_source_map(extracted_path, case_id)
    build_workbook(out_path, new_fields, legacy_catalog, case_source_map)

    print(
        json.dumps(
            {
                "ok": True,
                "generatedAt": datetime.now(timezone.utc).isoformat(),
                "caseId": case_id,
                "out": str(out_path),
                "newFieldOptions": len(new_fields),
                "legacyFiles": len(legacy_catalog),
                "legacyRows": sum(len(v) for v in legacy_catalog.values()),
                "caseSourceValues": len(case_source_map),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
